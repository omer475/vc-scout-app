import asyncio
import io
import json
import os
from datetime import datetime
from typing import Optional
from pathlib import Path

# Load .env
env_path = Path(__file__).parent / ".env"
if env_path.exists():
    for line in env_path.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            key, val = line.split("=", 1)
            os.environ.setdefault(key.strip(), val.strip())

from fastapi import FastAPI, Depends, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import or_
from openpyxl import Workbook

from database import init_db, get_db, SessionLocal, Source, Topic, Company, ScanLog, company_topics
from scraper import crawl_and_extract

app = FastAPI(title="VC Scout", version="2.0.0")

_cors_env = os.environ.get("CORS_ORIGINS", "*").strip()
_cors_origins = ["*"] if _cors_env == "*" else [o.strip() for o in _cors_env.split(",") if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_cors_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

init_db()


# ── Schemas ──

class SourceCreate(BaseModel):
    name: str
    url: str

class SourceOut(BaseModel):
    id: int
    name: str
    url: str
    is_active: bool
    created_at: datetime
    last_scraped_at: Optional[datetime] = None
    class Config:
        from_attributes = True

class TopicCreate(BaseModel):
    name: str

class TopicOut(BaseModel):
    id: int
    name: str
    is_active: bool
    class Config:
        from_attributes = True

class CompanyOut(BaseModel):
    id: int
    name: str
    description: Optional[str] = None
    website: Optional[str] = None
    source_url: Optional[str] = None
    source_name: Optional[str] = None
    page_url: Optional[str] = None
    industry: Optional[str] = None
    location: Optional[str] = None
    founded_year: Optional[int] = None
    founders: list[str] = []
    funding_stage: Optional[str] = None
    seeking_amount: Optional[str] = None
    is_raising: bool = False
    activity_type: Optional[str] = None
    raising_evidence: Optional[str] = None
    is_seen: bool
    is_new: bool
    discovered_at: datetime
    class Config:
        from_attributes = True

    @classmethod
    def from_orm_company(cls, c: Company) -> "CompanyOut":
        try:
            founders = json.loads(c.founders) if c.founders else []
            if not isinstance(founders, list):
                founders = []
        except Exception:
            founders = []
        return cls(
            id=c.id,
            name=c.name,
            description=c.description,
            website=c.website,
            source_url=c.source_url,
            source_name=c.source_name,
            page_url=c.page_url,
            industry=c.industry,
            location=c.location,
            founded_year=c.founded_year,
            founders=founders,
            funding_stage=c.funding_stage,
            seeking_amount=c.seeking_amount,
            is_raising=bool(c.is_raising),
            activity_type=c.activity_type,
            raising_evidence=c.raising_evidence,
            is_seen=bool(c.is_seen),
            is_new=bool(c.is_new),
            discovered_at=c.discovered_at,
        )

class ScanConfig(BaseModel):
    country: Optional[str] = None
    topic_ids: Optional[list[int]] = None

class ScanStart(BaseModel):
    scan_id: int
    status: str
    sources_total: int

class ScanStatus(BaseModel):
    scan_id: int
    status: str
    sources_scanned: int
    sources_total: int
    new_companies_found: int
    pages_crawled: int
    started_at: datetime
    finished_at: Optional[datetime] = None

class DashboardStats(BaseModel):
    total_companies: int
    new_companies: int
    raising_companies: int
    raising_now: int = 0
    recent_round: int = 0
    vc_portfolio: int = 0
    demo_day: int = 0
    total_sources: int
    active_topics: int
    last_scan: Optional[datetime] = None


# ── Sources ──

@app.get("/api/sources", response_model=list[SourceOut])
def list_sources(db: Session = Depends(get_db)):
    return db.query(Source).order_by(Source.created_at.desc()).all()

@app.post("/api/sources", response_model=SourceOut)
def add_source(data: SourceCreate, db: Session = Depends(get_db)):
    existing = db.query(Source).filter(Source.url == data.url).first()
    if existing:
        raise HTTPException(400, "This URL is already added")
    source = Source(name=data.name, url=data.url)
    db.add(source)
    db.commit()
    db.refresh(source)
    return source

@app.delete("/api/sources/{source_id}")
def delete_source(source_id: int, db: Session = Depends(get_db)):
    source = db.query(Source).filter(Source.id == source_id).first()
    if not source:
        raise HTTPException(404, "Source not found")
    db.delete(source)
    db.commit()
    return {"ok": True}

@app.patch("/api/sources/{source_id}/toggle")
def toggle_source(source_id: int, db: Session = Depends(get_db)):
    source = db.query(Source).filter(Source.id == source_id).first()
    if not source:
        raise HTTPException(404, "Source not found")
    source.is_active = not source.is_active
    db.commit()
    return {"ok": True, "is_active": source.is_active}


# ── Topics ──

@app.get("/api/topics", response_model=list[TopicOut])
def list_topics(db: Session = Depends(get_db)):
    return db.query(Topic).order_by(Topic.name).all()

@app.post("/api/topics", response_model=TopicOut)
def add_topic(data: TopicCreate, db: Session = Depends(get_db)):
    existing = db.query(Topic).filter(Topic.name == data.name).first()
    if existing:
        raise HTTPException(400, "Topic already exists")
    topic = Topic(name=data.name)
    db.add(topic)
    db.commit()
    db.refresh(topic)
    return topic

@app.delete("/api/topics/{topic_id}")
def delete_topic(topic_id: int, db: Session = Depends(get_db)):
    topic = db.query(Topic).filter(Topic.id == topic_id).first()
    if not topic:
        raise HTTPException(404, "Topic not found")
    db.delete(topic)
    db.commit()
    return {"ok": True}

@app.patch("/api/topics/{topic_id}/toggle")
def toggle_topic(topic_id: int, db: Session = Depends(get_db)):
    topic = db.query(Topic).filter(Topic.id == topic_id).first()
    if not topic:
        raise HTTPException(404, "Topic not found")
    topic.is_active = not topic.is_active
    db.commit()
    return {"ok": True, "is_active": topic.is_active}


# ── Scan ──

PER_SOURCE_TIMEOUT_SECONDS = 180


async def _scan_one_source(source: Source, country: Optional[str], active_topics: list[Topic], db: Session, batch_names: set[str]) -> int:
    """Scrape one source and persist matching companies. Returns (pages_crawled, new_company_count)."""
    import re as _re
    blob = (source.name + " " + source.url).lower()
    if any(kw in blob for kw in ["crowdfunding", "startupfon", "fonbulucu", "fongogo", "seedblink", "arikovani", "arıkovan"]):
        mode = "crowdfunding"
    elif any(kw in blob for kw in ["vc portfolio", "vc ", "212.vc", "revo.vc", "collectivespark", "earlybird", "logo ventures", "maki", "diffusion"]):
        mode = "vc_portfolio"
    elif any(kw in blob for kw in ["webrazzi", "yatirim-turu", "yatırım-turu", "girisim-haberleri", "girisimhaber", "startupsmagazine", "news"]):
        mode = "news"
    elif any(kw in blob for kw in ["cekirdek", "çekirdek", "kworks", "demo-day", "demo day", "cohort", "accelerator", "techstars"]):
        mode = "demo_day"
    else:
        mode = "default"
    force_browser = any(kw in blob for kw in ["startupfon", "fonbulucu", "seedblink", "arikovani", "arıkovan", "212.vc", "revo.vc", "collectivespark"])

    scraped, pages = await crawl_and_extract(
        url=source.url,
        source_name=source.name,
        topics=None,
        country=country or None,
        source_mode=mode,
        force_browser=force_browser,
    )
    source.last_scraped_at = datetime.utcnow()

    new_count = 0
    for sc in scraped:
        if not sc.is_raising:
            continue
        if country:
            loc_lower = (sc.location or "").lower()
            country_lower = country.lower()
            TURKEY_TOKENS = ["turkey", "türkiye", "turkiye", "istanbul", "ankara", "izmir", "bursa", "antalya", "eskisehir", "eskişehir", "kocaeli"]
            is_turkey_req = country_lower in ("turkey", "türkiye", "turkiye")
            if sc.activity_type == "recent_round":
                if is_turkey_req:
                    if not loc_lower or not any(t in loc_lower for t in TURKEY_TOKENS):
                        continue
                elif not loc_lower or country_lower not in loc_lower:
                    continue
            else:
                if loc_lower and country_lower not in loc_lower and not (is_turkey_req and any(t in loc_lower for t in TURKEY_TOKENS)):
                    continue

        norm_name = _re.sub(r'[^a-z0-9]', '', sc.name.lower())
        if norm_name in batch_names:
            continue
        is_dup = False
        for existing in db.query(Company).all():
            if _re.sub(r'[^a-z0-9]', '', existing.name.lower()) == norm_name:
                is_dup = True
                break
        if is_dup:
            continue
        batch_names.add(norm_name)

        company = Company(
            name=sc.name,
            description=sc.description,
            website=sc.website,
            source_url=sc.source_url,
            source_name=sc.source_name,
            page_url=sc.page_url,
            industry=sc.industry,
            location=sc.location,
            founded_year=sc.founded_year,
            founders=json.dumps(sc.founders) if sc.founders else None,
            funding_stage=sc.funding_stage,
            seeking_amount=sc.seeking_amount,
            is_raising=bool(sc.is_raising),
            activity_type=sc.activity_type,
            raising_evidence=sc.raising_evidence,
            is_new=True,
            is_seen=False,
        )

        searchable = f"{sc.name} {sc.description} {sc.industry}".lower()
        for topic in active_topics:
            pattern = r'\b' + _re.escape(topic.name.lower()) + r'\b'
            if _re.search(pattern, searchable):
                company.topics.append(topic)

        db.add(company)
        new_count += 1

    return pages, new_count


async def _run_scan_background(scan_id: int, country: Optional[str], topic_ids: Optional[list[int]]):
    db = SessionLocal()
    try:
        scan = db.query(ScanLog).filter(ScanLog.id == scan_id).first()
        if not scan:
            return

        sources = db.query(Source).filter(Source.is_active == True).all()
        if topic_ids:
            active_topics = db.query(Topic).filter(Topic.id.in_(topic_ids)).all()
        else:
            active_topics = db.query(Topic).filter(Topic.is_active == True).all()

        batch_names: set[str] = set()
        for source in sources:
            try:
                pages, new_count = await asyncio.wait_for(
                    _scan_one_source(source, country, active_topics, db, batch_names),
                    timeout=PER_SOURCE_TIMEOUT_SECONDS,
                )
                scan.sources_scanned = (scan.sources_scanned or 0) + 1
                scan.pages_crawled = (scan.pages_crawled or 0) + pages
                scan.new_companies_found = (scan.new_companies_found or 0) + new_count
                db.commit()
            except asyncio.TimeoutError:
                print(f"[scan {scan_id}] source {source.url} timed out after {PER_SOURCE_TIMEOUT_SECONDS}s")
                db.rollback()
            except Exception as e:
                print(f"[scan {scan_id}] source {source.url} errored: {e}")
                db.rollback()

        scan.finished_at = datetime.utcnow()
        scan.status = "completed"
        db.commit()
    except Exception as e:
        print(f"[scan {scan_id}] fatal: {e}")
        try:
            scan.status = "failed"
            scan.finished_at = datetime.utcnow()
            db.commit()
        except Exception:
            pass
    finally:
        db.close()


@app.post("/api/scan", response_model=ScanStart)
async def run_scan(config: ScanConfig = ScanConfig(), db: Session = Depends(get_db)):
    sources = db.query(Source).filter(Source.is_active == True).all()
    if not sources:
        raise HTTPException(400, "No active sources. Add sources first.")

    scan = ScanLog(status="running")
    db.add(scan)
    db.commit()
    db.refresh(scan)

    asyncio.create_task(_run_scan_background(scan.id, config.country, config.topic_ids))

    return ScanStart(scan_id=scan.id, status="running", sources_total=len(sources))


@app.get("/api/scan/{scan_id}", response_model=ScanStatus)
def get_scan_status(scan_id: int, db: Session = Depends(get_db)):
    scan = db.query(ScanLog).filter(ScanLog.id == scan_id).first()
    if not scan:
        raise HTTPException(404, "Scan not found")
    sources_total = db.query(Source).filter(Source.is_active == True).count()
    return ScanStatus(
        scan_id=scan.id,
        status=scan.status,
        sources_scanned=scan.sources_scanned or 0,
        sources_total=sources_total,
        new_companies_found=scan.new_companies_found or 0,
        pages_crawled=scan.pages_crawled or 0,
        started_at=scan.started_at,
        finished_at=scan.finished_at,
    )


# ── Companies ──

@app.get("/api/companies", response_model=list[CompanyOut])
def list_companies(
    new_only: bool = Query(False),
    raising_only: bool = Query(False),
    activity_type: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    topic_filter: bool = Query(False),
    year_min: Optional[int] = Query(None),
    year_max: Optional[int] = Query(None),
    location: Optional[str] = Query(None),
    db: Session = Depends(get_db),
):
    q = db.query(Company).order_by(Company.discovered_at.desc())
    if new_only:
        q = q.filter(Company.is_new == True)
    if raising_only:
        q = q.filter(Company.is_raising == True)
    if activity_type:
        q = q.filter(Company.activity_type == activity_type)
    if year_min:
        q = q.filter(Company.founded_year >= year_min)
    if year_max:
        q = q.filter(Company.founded_year <= year_max)
    if search:
        q = q.filter(
            or_(
                Company.name.ilike(f"%{search}%"),
                Company.description.ilike(f"%{search}%"),
                Company.industry.ilike(f"%{search}%"),
                Company.founders.ilike(f"%{search}%"),
            )
        )
    if location:
        q = q.filter(Company.location.ilike(f"%{location}%"))
    if topic_filter:
        active_topics = db.query(Topic).filter(Topic.is_active == True).all()
        if active_topics:
            q = q.filter(Company.topics.any(Topic.id.in_([t.id for t in active_topics])))
    return [CompanyOut.from_orm_company(c) for c in q.all()]

@app.patch("/api/companies/{company_id}/mark-seen")
def mark_seen(company_id: int, db: Session = Depends(get_db)):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(404)
    company.is_seen = True
    company.is_new = False
    db.commit()
    return {"ok": True}

@app.patch("/api/companies/mark-all-seen")
def mark_all_seen(db: Session = Depends(get_db)):
    db.query(Company).filter(Company.is_new == True).update({"is_new": False, "is_seen": True})
    db.commit()
    return {"ok": True}

@app.delete("/api/companies/{company_id}")
def delete_company(company_id: int, db: Session = Depends(get_db)):
    company = db.query(Company).filter(Company.id == company_id).first()
    if not company:
        raise HTTPException(404)
    db.delete(company)
    db.commit()
    return {"ok": True}


# ── Export ──

@app.get("/api/export/excel")
def export_excel(new_only: bool = Query(False), raising_only: bool = Query(False), db: Session = Depends(get_db)):
    q = db.query(Company).order_by(Company.discovered_at.desc())
    if new_only:
        q = q.filter(Company.is_new == True)
    if raising_only:
        q = q.filter(Company.is_raising == True)
    companies = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Companies"
    headers = ["Name", "Description", "Founders", "Raising?", "Stage", "Seeking", "Evidence", "Website", "Industry", "Location", "Founded", "Source", "Found On Page", "Discovered", "Status"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = ws.cell(row=1, column=col_idx).font.copy(bold=True)

    for c in companies:
        try:
            founders_list = json.loads(c.founders) if c.founders else []
        except Exception:
            founders_list = []
        ws.append([
            c.name, c.description or "",
            ", ".join(founders_list),
            "Yes" if c.is_raising else "",
            c.funding_stage or "",
            c.seeking_amount or "",
            c.raising_evidence or "",
            c.website or "", c.industry or "",
            c.location or "", c.founded_year or "",
            c.source_name or "", c.page_url or "",
            c.discovered_at.strftime("%Y-%m-%d %H:%M") if c.discovered_at else "",
            "New" if c.is_new else "Seen",
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"vc_scout_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ── Dashboard ──

@app.get("/api/dashboard", response_model=DashboardStats)
def dashboard_stats(db: Session = Depends(get_db)):
    last_scan = db.query(ScanLog).order_by(ScanLog.started_at.desc()).first()
    return DashboardStats(
        total_companies=db.query(Company).count(),
        new_companies=db.query(Company).filter(Company.is_new == True).count(),
        raising_companies=db.query(Company).filter(Company.is_raising == True).count(),
        raising_now=db.query(Company).filter(Company.activity_type == "raising").count(),
        recent_round=db.query(Company).filter(Company.activity_type == "recent_round").count(),
        vc_portfolio=db.query(Company).filter(Company.activity_type == "vc_portfolio").count(),
        demo_day=db.query(Company).filter(Company.activity_type == "demo_day").count(),
        total_sources=db.query(Source).count(),
        active_topics=db.query(Topic).filter(Topic.is_active == True).count(),
        last_scan=last_scan.started_at if last_scan else None,
    )


# ── Seed ──

@app.post("/api/seed")
def seed_topics(db: Session = Depends(get_db)):
    defaults = [
        "Fintech", "SaaS", "AI", "Machine Learning", "HealthTech",
        "EdTech", "E-commerce", "Blockchain", "Cybersecurity", "IoT",
        "CleanTech", "DeepTech", "Biotech", "Robotics", "Gaming",
        "Logistics", "PropTech", "InsurTech", "FoodTech", "AgriTech",
    ]
    added = 0
    for name in defaults:
        if not db.query(Topic).filter(Topic.name == name).first():
            db.add(Topic(name=name))
            added += 1
    db.commit()
    return {"ok": True, "topics_added": added}


@app.post("/api/seed-companies")
def seed_companies(db: Session = Depends(get_db)):
    """Seed well-known Turkish startups that should be in any professional database."""
    import re as _re
    TURKISH_STARTUPS = [
        {"name": "Getir", "description": "Ultra-fast grocery and essentials delivery platform, pioneering the quick-commerce model globally. Delivers orders in minutes from dark stores.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://getir.com", "founded_year": 2015},
        {"name": "Trendyol", "description": "Turkey's largest e-commerce platform with fashion, electronics, groceries and more. Part of Alibaba Group.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://www.trendyol.com", "founded_year": 2010},
        {"name": "Peak Games", "description": "Mobile gaming studio known for hit casual games including Toy Blast and Toon Blast. Acquired by Zynga for $1.8B.", "industry": "Gaming", "location": "Istanbul, Turkey", "website": "https://peak.com", "founded_year": 2010},
        {"name": "Dream Games", "description": "Mobile gaming studio behind Royal Match, one of the highest-grossing puzzle games globally. Valued at $2.75B.", "industry": "Gaming", "location": "Istanbul, Turkey", "website": "https://www.dreamgames.com", "founded_year": 2019},
        {"name": "Hepsiburada", "description": "One of Turkey's leading e-commerce platforms offering electronics, fashion, home goods and more. Listed on NASDAQ.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://www.hepsiburada.com", "founded_year": 2000},
        {"name": "Yemeksepeti", "description": "Turkey's largest online food ordering and delivery platform, connecting restaurants with consumers. Acquired by Delivery Hero.", "industry": "FoodTech", "location": "Istanbul, Turkey", "website": "https://www.yemeksepeti.com", "founded_year": 2001},
        {"name": "Insider", "description": "AI-native platform for individualized, cross-channel customer experiences. Serves 1200+ brands in 26 countries.", "industry": "MarTech", "location": "Istanbul, Turkey", "website": "https://useinsider.com", "founded_year": 2012},
        {"name": "Param", "description": "Digital payment and financial technology platform providing payment gateway, wallet and lending services.", "industry": "Fintech", "location": "Istanbul, Turkey", "website": "https://param.com.tr", "founded_year": 2015},
        {"name": "iyzico", "description": "Leading Turkish online payment platform enabling businesses to accept credit cards and alternative payment methods. Acquired by PayU.", "industry": "Fintech", "location": "Istanbul, Turkey", "website": "https://www.iyzico.com", "founded_year": 2013},
        {"name": "Modanisa", "description": "Global modest fashion e-commerce platform serving customers in 140+ countries with a curated selection of modest clothing brands.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://www.modanisa.com", "founded_year": 2011},
        {"name": "Armut", "description": "Turkey's largest local services marketplace connecting consumers with service professionals. Acquired by Delivery Hero's Yemeksepeti.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://armut.com", "founded_year": 2011},
        {"name": "Sahibinden.com", "description": "Turkey's largest classified ads platform for real estate, vehicles, and second-hand goods with millions of monthly visitors.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://www.sahibinden.com", "founded_year": 2000},
        {"name": "n11.com", "description": "Major Turkish online marketplace offering a wide range of products from electronics to fashion, backed by SK Group and Dogus Group.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://www.n11.com", "founded_year": 2013},
        {"name": "Letgo", "description": "Mobile-first marketplace for buying and selling secondhand goods locally. Merged with OfferUp.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://www.letgo.com", "founded_year": 2015},
        {"name": "Scotty", "description": "On-demand motorcycle taxi and courier service platform operating in Istanbul.", "industry": "Mobility", "location": "Istanbul, Turkey", "website": "https://scotty.app", "founded_year": 2019},
        {"name": "BiTaksi", "description": "Turkey's leading taxi-hailing mobile application, connecting passengers with licensed taxi drivers.", "industry": "Mobility", "location": "Istanbul, Turkey", "website": "https://www.bitaksi.com", "founded_year": 2013},
        {"name": "Marti", "description": "Turkey's leading micro-mobility provider operating e-scooters, e-mopeds and e-bikes. Listed on NYSE.", "industry": "Mobility", "location": "Istanbul, Turkey", "website": "https://www.marti.tech", "founded_year": 2018},
        {"name": "Udemy", "description": "Global online learning marketplace with 200M+ students. Founded by Turkish entrepreneur Eren Bali.", "industry": "EdTech", "location": "Istanbul, Turkey", "website": "https://www.udemy.com", "founded_year": 2010},
        {"name": "GittiGidiyor", "description": "One of Turkey's oldest and largest online auction and shopping platforms, an eBay company.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://www.gittigidiyor.com", "founded_year": 2001},
        {"name": "Gram Games", "description": "Mobile gaming studio known for Merge Magic and 1010! games. Acquired by Zynga.", "industry": "Gaming", "location": "Istanbul, Turkey", "website": "https://gram.gs", "founded_year": 2012},
        {"name": "Spyke Games", "description": "Mobile gaming startup building social competitive games, founded by former Peak Games executives.", "industry": "Gaming", "location": "Istanbul, Turkey", "website": "https://www.spykegames.com", "founded_year": 2021},
        {"name": "Ace Games", "description": "Mobile gaming studio developing casual and hyper-casual games for global markets.", "industry": "Gaming", "location": "Istanbul, Turkey", "website": "https://www.ace.games", "founded_year": 2019},
        {"name": "Masomo", "description": "Mobile gaming company known for its sports games including Head Ball 2 with 100M+ downloads.", "industry": "Gaming", "location": "Istanbul, Turkey", "website": "https://masomo.com", "founded_year": 2012},
        {"name": "Foriba", "description": "Leading e-invoicing and e-transformation platform in Turkey, serving 50,000+ businesses with digital tax compliance.", "industry": "Fintech", "location": "Istanbul, Turkey", "website": "https://www.foriba.com", "founded_year": 2009},
        {"name": "Papara", "description": "Digital banking and payments platform offering instant money transfers, virtual cards, and cashback rewards.", "industry": "Fintech", "location": "Istanbul, Turkey", "website": "https://www.papara.com", "founded_year": 2016},
        {"name": "Colendi", "description": "AI-powered embedded finance platform providing credit scoring, BNPL, and digital banking infrastructure.", "industry": "Fintech", "location": "Istanbul, Turkey", "website": "https://www.colendi.com", "founded_year": 2018},
        {"name": "Figopara", "description": "Supply chain finance platform enabling SMEs to access early payment on their invoices through reverse factoring.", "industry": "Fintech", "location": "Istanbul, Turkey", "website": "https://figopara.com", "founded_year": 2017},
        {"name": "Paraşüt", "description": "Cloud-based accounting and financial management SaaS platform for SMEs in Turkey.", "industry": "Fintech", "location": "Istanbul, Turkey", "website": "https://www.parasut.com", "founded_year": 2013},
        {"name": "Craftbase", "description": "AI-powered inventory and production management platform for handmade and craft businesses.", "industry": "SaaS", "location": "Istanbul, Turkey", "website": "https://craftbase.app", "founded_year": 2020},
        {"name": "Picus Security", "description": "Breach and attack simulation platform that continuously validates security controls and measures cyber risk.", "industry": "Cybersecurity", "location": "Ankara, Turkey", "website": "https://www.picussecurity.com", "founded_year": 2013},
        {"name": "SOCRadar", "description": "Extended Threat Intelligence platform combining attack surface management, digital risk protection and threat intelligence.", "industry": "Cybersecurity", "location": "Istanbul, Turkey", "website": "https://socradar.io", "founded_year": 2018},
        {"name": "Upstash", "description": "Serverless data platform offering Redis-compatible and Kafka-compatible databases for modern cloud applications.", "industry": "DevTools", "location": "Istanbul, Turkey", "website": "https://upstash.com", "founded_year": 2021},
        {"name": "Cbot", "description": "Enterprise conversational AI platform providing chatbot and virtual assistant solutions for customer service.", "industry": "AI", "location": "Istanbul, Turkey", "website": "https://www.cbot.ai", "founded_year": 2015},
        {"name": "UserGuiding", "description": "No-code user onboarding and product adoption platform helping SaaS companies create interactive product tours.", "industry": "SaaS", "location": "Istanbul, Turkey", "website": "https://userguiding.com", "founded_year": 2017},
        {"name": "Storyly", "description": "Interactive content platform enabling mobile apps and websites to create Instagram-like stories for user engagement.", "industry": "MarTech", "location": "Istanbul, Turkey", "website": "https://storyly.io", "founded_year": 2019},
        {"name": "Segmentify", "description": "AI-powered personalization and customer engagement platform for e-commerce businesses.", "industry": "MarTech", "location": "Istanbul, Turkey", "website": "https://www.segmentify.com", "founded_year": 2015},
        {"name": "Moova", "description": "Last-mile delivery technology platform using crowdsourced drivers for same-day and next-day deliveries.", "industry": "Logistics", "location": "Istanbul, Turkey", "website": "https://moova.io", "founded_year": 2019},
        {"name": "Cimri", "description": "Turkey's leading price comparison platform helping consumers find the best deals across online retailers.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://www.cimri.com", "founded_year": 2008},
        {"name": "Kolay IK", "description": "Cloud-based HR management platform for SMEs offering payroll, time tracking, and employee management.", "industry": "HRTech", "location": "Istanbul, Turkey", "website": "https://www.kolayik.com", "founded_year": 2015},
        {"name": "Otokoç", "description": "Turkey's largest fleet leasing and car rental company with digital fleet management solutions.", "industry": "Mobility", "location": "Istanbul, Turkey", "website": "https://www.otokoc.com.tr", "founded_year": 1928},
        {"name": "Lidio", "description": "Buy Now Pay Later platform enabling installment payments for online and offline purchases in Turkey.", "industry": "Fintech", "location": "Istanbul, Turkey", "website": "https://lidio.com", "founded_year": 2020},
        {"name": "Iyzi", "description": "Smart payment solutions and POS systems for small businesses and merchants in Turkey.", "industry": "Fintech", "location": "Istanbul, Turkey", "website": "https://iyzi.co", "founded_year": 2019},
        {"name": "Brisa", "description": "Turkey's leading tire manufacturer with smart tire technology and connected fleet solutions.", "industry": "DeepTech", "location": "Istanbul, Turkey", "website": "https://www.brisa.com.tr", "founded_year": 1974},
        {"name": "Aposto", "description": "Newsletter and content subscription platform for Turkish-language journalism and media.", "industry": "Media", "location": "Istanbul, Turkey", "website": "https://aposto.com", "founded_year": 2020},
        {"name": "Mundi", "description": "B2B cross-border trade finance platform providing working capital solutions for exporters.", "industry": "Fintech", "location": "Istanbul, Turkey", "website": "https://www.mundi.io", "founded_year": 2020},
        {"name": "Akinon", "description": "Headless commerce platform providing enterprise e-commerce infrastructure and omnichannel solutions.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://akinon.com", "founded_year": 2014},
        {"name": "Karaca", "description": "Turkish home and kitchenware brand with D2C e-commerce platform and smart home products.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://www.karaca.com", "founded_year": 1973},
        {"name": "HotelRunner", "description": "Cloud-based hotel distribution and channel management platform connecting hotels to 200+ OTAs and booking channels.", "industry": "Travel", "location": "Istanbul, Turkey", "website": "https://www.hotelrunner.com", "founded_year": 2012},
        {"name": "Dolap", "description": "Turkey's largest secondhand fashion marketplace enabling users to buy and sell pre-owned clothing.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://dolap.com", "founded_year": 2015},
        {"name": "Tarfin", "description": "Agricultural fintech platform providing farmers with access to quality farm inputs through digital credit solutions.", "industry": "AgriTech", "location": "Istanbul, Turkey", "website": "https://tarfin.com", "founded_year": 2017},
        {"name": "Vivense", "description": "Online furniture and home living marketplace offering a wide range of furniture and home decor products.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://www.vivense.com", "founded_year": 2016},
        {"name": "TabbyPay", "description": "Payments infrastructure and BNPL solutions provider for the MENA and Turkey region.", "industry": "Fintech", "location": "Istanbul, Turkey", "website": "https://tabby.ai", "founded_year": 2019},
        {"name": "Sendeo", "description": "Technology-driven logistics company providing e-commerce fulfillment and last-mile delivery services in Turkey.", "industry": "Logistics", "location": "Istanbul, Turkey", "website": "https://www.sendeo.com.tr", "founded_year": 2017},
        {"name": "Baykar", "description": "Defense technology company known for developing TB2 Bayraktar armed drones and Akinci UAVs.", "industry": "DeepTech", "location": "Istanbul, Turkey", "website": "https://www.baykartech.com", "founded_year": 1984},
        {"name": "HAVELSAN", "description": "Defense and IT company developing software-intensive systems for military and civilian applications.", "industry": "DeepTech", "location": "Ankara, Turkey", "website": "https://www.havelsan.com.tr", "founded_year": 1982},
        {"name": "STM", "description": "Defense technology company providing cyber security, naval platform integration, and satellite systems.", "industry": "Cybersecurity", "location": "Ankara, Turkey", "website": "https://www.stm.com.tr", "founded_year": 1991},
        {"name": "Roketsan", "description": "Leading Turkish defense company specializing in rockets, missiles, and munitions systems.", "industry": "DeepTech", "location": "Ankara, Turkey", "website": "https://www.roketsan.com.tr", "founded_year": 1988},
        {"name": "Logo Yazılım", "description": "Turkey's largest enterprise software company providing ERP, HR, and business management solutions.", "industry": "SaaS", "location": "Istanbul, Turkey", "website": "https://www.logo.com.tr", "founded_year": 1984},
        {"name": "Ebebek", "description": "Turkey's largest baby and kids products retailer with omnichannel presence and e-commerce platform.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://www.ebebek.com", "founded_year": 2004},
        {"name": "Obilet", "description": "Online bus, flight and ferry ticket booking platform, Turkey's leading ground transportation marketplace.", "industry": "Travel", "location": "Istanbul, Turkey", "website": "https://www.obilet.com", "founded_year": 2014},
        {"name": "Enuygun", "description": "Travel price comparison and booking platform for flights, hotels, car rentals and insurance in Turkey.", "industry": "Travel", "location": "Istanbul, Turkey", "website": "https://www.enuygun.com", "founded_year": 2008},
        {"name": "Solvoyo", "description": "End-to-end supply chain planning and analytics platform using AI for demand forecasting and inventory optimization.", "industry": "SaaS", "location": "Istanbul, Turkey", "website": "https://www.solvoyo.com", "founded_year": 2013},
        {"name": "Jotform", "description": "Online form builder with 25M+ users globally. Founded by Turkish entrepreneur Aytekin Tank.", "industry": "SaaS", "location": "Istanbul, Turkey", "website": "https://www.jotform.com", "founded_year": 2006},
        {"name": "Gratis", "description": "Turkey's largest beauty and personal care retail chain with e-commerce platform.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://www.grfratis.com", "founded_year": 2008},
        {"name": "Ticimax", "description": "E-commerce infrastructure platform providing turnkey online store solutions for Turkish businesses.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://www.ticimax.com", "founded_year": 2006},
        {"name": "Prisync", "description": "Competitive pricing intelligence and dynamic pricing software for e-commerce businesses globally.", "industry": "SaaS", "location": "Istanbul, Turkey", "website": "https://prisync.com", "founded_year": 2013},
        {"name": "Bulutfon", "description": "Cloud communication platform providing virtual PBX, call center, and business phone solutions.", "industry": "SaaS", "location": "Istanbul, Turkey", "website": "https://www.bulutfon.com", "founded_year": 2013},
        {"name": "Vispera", "description": "AI-powered image recognition platform for retail shelf analytics and in-store execution monitoring.", "industry": "AI", "location": "Istanbul, Turkey", "website": "https://www.vispera.co", "founded_year": 2014},
        {"name": "Mobiroller", "description": "No-code mobile app builder enabling businesses to create iOS and Android apps without programming.", "industry": "SaaS", "location": "Istanbul, Turkey", "website": "https://www.mobiroller.com", "founded_year": 2012},
        {"name": "Sertifier", "description": "Digital credential and certificate management platform for education and corporate training.", "industry": "EdTech", "location": "Istanbul, Turkey", "website": "https://sertifier.com", "founded_year": 2018},
        {"name": "Perceptive Analytics", "description": "AI-powered sales intelligence platform providing B2B data enrichment and lead scoring.", "industry": "AI", "location": "Istanbul, Turkey", "website": "https://perceptiveanalytics.co", "founded_year": 2020},
        {"name": "Bionluk", "description": "Turkey's largest freelancer marketplace connecting businesses with freelance professionals.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://bionluk.com", "founded_year": 2013},
        {"name": "Ikas", "description": "E-commerce platform enabling merchants to build online stores with integrated payment and shipping.", "industry": "E-commerce", "location": "Istanbul, Turkey", "website": "https://ikas.com", "founded_year": 2017},
        {"name": "Hubtic", "description": "Marketing automation and CRM platform for Turkish SMEs with email marketing and customer tracking.", "industry": "MarTech", "location": "Istanbul, Turkey", "website": "https://hubtic.com", "founded_year": 2019},
        {"name": "Midas", "description": "Social trading and stock investment app making capital markets accessible to retail investors in Turkey.", "industry": "Fintech", "location": "Istanbul, Turkey", "website": "https://www.getmidas.com", "founded_year": 2021},
        {"name": "BulutMD", "description": "Cloud-based hospital information management system and clinical decision support software.", "industry": "HealthTech", "location": "Istanbul, Turkey", "website": "https://www.bulutmd.com", "founded_year": 2015},
        {"name": "PayTR", "description": "Payment technology company providing virtual POS, payment gateway and marketplace payment solutions.", "industry": "Fintech", "location": "Istanbul, Turkey", "website": "https://www.paytr.com", "founded_year": 2015},
        {"name": "Tapu.com", "description": "Turkey's digital real estate platform providing property search, valuation and transaction services.", "industry": "PropTech", "location": "Istanbul, Turkey", "website": "https://www.tapu.com", "founded_year": 2014},
        {"name": "Emphy", "description": "AI-powered HR platform automating recruitment, employee engagement and workforce analytics.", "industry": "HRTech", "location": "Istanbul, Turkey", "website": "https://emphy.co", "founded_year": 2020},
    ]

    added = 0
    for s in TURKISH_STARTUPS:
        norm = _re.sub(r'[^a-z0-9]', '', s["name"].lower())
        exists = False
        for c in db.query(Company).all():
            if _re.sub(r'[^a-z0-9]', '', c.name.lower()) == norm:
                exists = True
                break
        if exists:
            continue
        company = Company(
            name=s["name"],
            description=s["description"],
            website=s.get("website", ""),
            industry=s["industry"],
            location=s["location"],
            founded_year=s.get("founded_year"),
            source_name="Curated",
            source_url="",
            page_url="",
            is_new=True,
            is_seen=False,
        )
        db.add(company)
        added += 1
    db.commit()
    return {"ok": True, "companies_added": added}


@app.post("/api/cleanup-db")
def cleanup_db(db: Session = Depends(get_db)):
    """Reclassify VC-portfolio entries to vc_portfolio activity_type and delete non-Turkish leakage."""
    reclassified = 0
    deleted = 0

    # Reclassify: entries from VC portfolio sources should have activity_type=vc_portfolio
    VC_SOURCE_MARKERS = ["212 vc", "revo capital", "collective spark", "earlybird", "logo ventures", "maki"]
    cs = db.query(Company).all()
    for c in cs:
        src = (c.source_name or "").lower()
        if any(m in src for m in VC_SOURCE_MARKERS) and c.activity_type != "vc_portfolio":
            c.activity_type = "vc_portfolio"
            reclassified += 1

    # Delete: companies whose description explicitly flags non-Turkish HQ
    NON_TR_MARKERS = [
        "uk-based", "u.k.-based", "us-based", "u.s.-based",
        "headquartered in london", "headquartered in new york",
        "headquartered in san francisco", "headquartered in berlin",
        "headquartered in amsterdam", "headquartered in dublin",
        "headquartered in tel aviv", "based in london", "based in new york",
        "based in the uk", "based in the us", "based in the united states",
        "based in the united kingdom",
    ]
    for c in db.query(Company).all():
        desc = (c.description or "").lower()
        if any(m in desc for m in NON_TR_MARKERS):
            db.delete(c)
            deleted += 1
    db.commit()
    return {"ok": True, "reclassified": reclassified, "deleted": deleted}


@app.post("/api/seed-sources-tr")
def seed_sources_tr(db: Session = Depends(get_db)):
    """Seed Turkish-focused sources that list startups actively seeking investment."""
    TURKISH_RAISING_SOURCES = [
        # Equity crowdfunding — currently raising
        {"name": "Startupfon (Equity Crowdfunding)", "url": "https://startupfon.com"},
        {"name": "Fonbulucu Invest (Equity Crowdfunding)", "url": "https://invest.fonbulucu.com"},
        {"name": "Arıkovanı (Equity Crowdfunding)", "url": "https://www.arikovani.com"},
        # News — recently raised rounds
        {"name": "Webrazzi Yatırım (News)", "url": "https://webrazzi.com/kategori/yatirim"},
        {"name": "Webrazzi Yatırım Etiket (News)", "url": "https://webrazzi.com/etiket/yatirim"},
        {"name": "GirişimHaber (News)", "url": "https://www.girisimhaber.com"},
        # Turkey-focused VC portfolio pages (recent rounds via VC)
        {"name": "212 VC Portfolio (News)", "url": "https://www.212.vc"},
        {"name": "Revo Capital Portfolio (News)", "url": "https://www.revo.vc"},
        {"name": "Collective Spark Portfolio (News)", "url": "https://collectivespark.com"},
        # Demo day / accelerator cohorts
        {"name": "ITU Çekirdek Cohort", "url": "https://itucekirdek.com/en/startups"},
        {"name": "KWORKS Cohort", "url": "https://kworks.ku.edu.tr/en/startups"},
    ]
    added = 0
    for s in TURKISH_RAISING_SOURCES:
        exists = db.query(Source).filter(Source.url == s["url"]).first()
        if exists:
            continue
        db.add(Source(name=s["name"], url=s["url"]))
        added += 1
    db.commit()
    return {"ok": True, "sources_added": added}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
